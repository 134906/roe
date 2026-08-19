import pandas as pd
import numpy as np
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
from .descriptives import get_descriptives
from openpyxl.formatting.rule import FormulaRule

# ---------- 内部辅助函数 ----------
def _write_num(cell, val, fmt=None):
    """写入数字，空值显示为空"""
    if val is None or pd.isna(val):
        cell.value = ""
        return
    cell.value = val
    if isinstance(val, (int, float)):
        if fmt is not None:
            cell.number_format = fmt
        else:
            if abs(val - round(val)) < 1e-12:
                cell.number_format = '0'
            else:
                cell.number_format = '0.00'

def _write_percent(cell, val):
    """写入百分比，空值显示为空"""
    if val is None or pd.isna(val):
        cell.value = ""
        return
    cell.value = val
    if isinstance(val, (int, float)):
        cell.number_format = '0.00%'

# ================================================================
# 1. Descriptives（描述统计）
# ================================================================
def write_descriptives_to_sheet(ws, desc_list):
    """写入描述统计表"""
    current_row = 1
    for title, df_desc in desc_list:
        if df_desc is not None and not df_desc.empty:
            ws.cell(row=current_row, column=1, value=title)
            current_row += 1
            headers = ['Variable', 'N', 'Min', 'Max', 'Mean', 'Std']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=h)
            current_row += 1
            for idx, row in df_desc.iterrows():
                ws.cell(row=current_row, column=1, value=idx)
                for col, field in enumerate(['N', 'Min', 'Max', 'Mean', 'Std'], start=2):
                    _write_num(ws.cell(row=current_row, column=col), row[field])
                current_row += 1
            current_row += 1
    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

# ================================================================
# 2. Sample Check（样本量核对）
# ================================================================
def write_sample_check_to_sheet(ws, kpi_check_df, channel_check_df, success, source=None):
    row = 1
    # 根据 source 确定样本量列名
    if source == 'data_check':
        sample_col_name = 'Data_Check样本量'
        empty_msg = "Data_Check 数据为空或未提供，已跳过样本量检查"
    else:
        sample_col_name = 'Full_Table样本量'
        empty_msg = "Full_Table 为空或未提供，已跳过样本量检查"

    # 如果数据为空，显示提示并返回
    if (kpi_check_df is None or kpi_check_df.empty) and (channel_check_df is None or channel_check_df.empty):
        ws.cell(row=row, column=1, value="样本量核对")
        row += 1
        ws.cell(row=row, column=1, value=empty_msg)
        return

    # ---- KPI 部分 ----
    ws.cell(row=row, column=1, value='KPI 样本量校对')
    row += 1
    headers = ['表名', 'KPI标签', sample_col_name, 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    if kpi_check_df is not None and not kpi_check_df.empty:
        for _, row_data in kpi_check_df.iterrows():
            ws.cell(row=row, column=1, value=row_data['TableName'])
            ws.cell(row=row, column=2, value=row_data['KPI_Label'])
            ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
            ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
            ws.cell(row=row, column=5, value=row_data['Diff'])
            row += 1
    row += 1

    # ---- 渠道部分 ----
    ws.cell(row=row, column=1, value='渠道样本量校对')
    row += 1
    headers = ['表名', '渠道名', sample_col_name, 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    if channel_check_df is not None and not channel_check_df.empty:
        for _, row_data in channel_check_df.iterrows():
            ws.cell(row=row, column=1, value=row_data['TableName'])
            ws.cell(row=row, column=2, value=row_data['Channel'])
            ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
            ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
            ws.cell(row=row, column=5, value=row_data['Diff'])
            row += 1
    row += 1

    ws.cell(row=row, column=1, value='样本量校对结果：')
    ws.cell(row=row, column=2, value='成功' if success else '失败（存在差异）')

    # ---- 添加条件格式到差异列（第5列） ----
    # 确定数据范围：KPI 和渠道的差异列都在第5列，行从第4行开始到当前行-1
    diff_start_row = 4  # 表头在第3行，数据从第4行开始
    diff_end_row = row - 2  # 最后一行数据（结果行之前）
    if diff_end_row >= diff_start_row:
        col_letter = get_column_letter(5)  # E列
        range_str = f'{col_letter}{diff_start_row}:{col_letter}{diff_end_row}'
        try:
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',   # 红
                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # 黄
                end_type='max', end_color='63BE7B'        # 绿
            )
            ws.conditional_formatting.add(range_str, rule)
        except Exception:
            pass

    # 调整列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)

# ================================================================
# 3. CFA Results（因子分析结果）
# ================================================================
def write_cfa_results_to_sheet(ws, fa_extra):
    """写入CFA结果（包括总方差、载荷矩阵、得分系数、SPSS语法等）"""
    var_id_map = fa_extra.get('var_id_map', {})
    var_labels = fa_extra.get('factor_var_labels', {})

    row = 1
    col_shift = 1
    # 头部信息
    ws.cell(row=row, column=col_shift, value="Spend time = 00:00:02")
    row += 1
    ws.cell(row=row, column=col_shift, value="Factor Result")
    row += 1
    now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    ws.cell(row=row, column=col_shift, value=f"TIME= {now}")
    row += 1
    rot_method = fa_extra.get('rotation', 'varimax').capitalize()
    ws.cell(row=row, column=col_shift, value=f"Rotation Method= {rot_method} Method")
    row += 3

    # 总方差解释
    ws.cell(row=row, column=col_shift, value="Total Variance Explained")
    row += 1
    ws.cell(row=row, column=col_shift, value="Initial Eigenvalues")
    row += 1
    ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
    row += 2
    eigenvals = fa_extra.get('eigenvals')
    var_ratio = fa_extra.get('var_ratio')
    cum_var = fa_extra.get('cum_var')
    if eigenvals is not None:
        ws.cell(row=row, column=col_shift, value="Component")
        ws.cell(row=row, column=col_shift+1, value="Total")
        ws.cell(row=row, column=col_shift+2, value="% of Variance")
        ws.cell(row=row, column=col_shift+3, value="Cumulative %")
        row += 1
        for i, (ev, vr, cv) in enumerate(zip(eigenvals, var_ratio, cum_var), start=1):
            ws.cell(row=row, column=col_shift, value=f"Factor{i}")
            _write_num(ws.cell(row=row, column=col_shift+1), ev)
            _write_num(ws.cell(row=row, column=col_shift+2), vr, fmt='0.00%')
            _write_num(ws.cell(row=row, column=col_shift+3), cv, fmt='0.00%')
            row += 1
        row += 1

    target = fa_extra.get('target')
    if target is None:
        return
    row_labels = target.index.tolist()
    col_labels = target.columns.tolist()
    n_factors = len(col_labels)
    sig_break = fa_extra.get('sig_break', 0.3)

    def write_first_three_cols(var_name, i, row_num):
        var_id = var_id_map.get(var_name, i+1)
        label = var_labels.get(var_name, var_name)
        ws.cell(row=row_num, column=col_shift, value=var_id)
        ws.cell(row=row_num, column=col_shift+1, value=var_name)
        ws.cell(row=row_num, column=col_shift+2, value=label)

    # 未旋转载荷
    unrot = fa_extra.get('unrot_loadings')
    if unrot is not None:
        ws.cell(row=row, column=col_shift, value="Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        for i, var_name in enumerate(row_labels):
            write_first_three_cols(var_name, i, row)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                _write_num(cell, unrot[i, j])
                if abs(unrot[i, j]) > sig_break:
                    cell.fill = yellow_fill
                    cell.font = bold_font
            row += 1
        row += 1

    # 旋转载荷
    rot = fa_extra.get('rot_loadings')
    if rot is not None:
        ws.cell(row=row, column=col_shift, value="Rotated Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Rotation Method: Procrustes with Kaiser Normalization.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, var_name in enumerate(row_labels):
            write_first_three_cols(var_name, i, row)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                _write_num(cell, rot[i, j])
                if target.iloc[i, j] != 0:
                    cell.font = Font(italic=True, color="FF0000")
                if abs(rot[i, j]) > sig_break:
                    if target.iloc[i, j] != 0:
                        cell.font = Font(bold=True, italic=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True)
                    cell.fill = yellow_fill
            row += 1
        row += 1

    # 规则矩阵
    ws.cell(row=row, column=col_shift, value="Available Rotation Rule Matrix")
    row += 1
    ws.cell(row=row, column=col_shift, value="You can Copy this Matrix to Rotation Rule Sheet.")
    row += 2
    ws.cell(row=row, column=col_shift, value="VariableID")
    ws.cell(row=row, column=col_shift+1, value="Variable")
    ws.cell(row=row, column=col_shift+2, value="Lable")
    for j, col_name in enumerate(col_labels):
        ws.cell(row=row, column=col_shift+3+j, value=col_name)
    row += 1
    for i, var_name in enumerate(row_labels):
        write_first_three_cols(var_name, i, row)
        for j in range(n_factors):
            val = rot[i, j] if rot is not None else 0
            if val > sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=1)
            elif val < -sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=-1)
            else:
                ws.cell(row=row, column=col_shift+3+j, value='')
        row += 1
    row += 1

    # 标准化得分系数
    score_std = fa_extra.get('score_coeff_std')
    if score_std is not None:
        ws.cell(row=row, column=col_shift, value="Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            write_first_three_cols(var_name, i, row)
            for j in range(n_factors):
                _write_num(ws.cell(row=row, column=col_shift+3+j), score_std[i, j])
            row += 1
        row += 1

    # 非标准化得分系数
    score_unstd_table = fa_extra.get('score_coeff_unstd_table')
    if score_unstd_table is not None:
        ws.cell(row=row, column=col_shift, value="Unstandardized Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            write_first_three_cols(var_name, i, row)
            for j in range(n_factors):
                _write_num(ws.cell(row=row, column=col_shift+3+j), score_unstd_table[i, j])
            row += 1
        row += 1

    # SPSS语法
    spss_syntax = fa_extra.get('spss_syntax')
    if spss_syntax:
        ws.cell(row=row, column=col_shift, value="SPSS Syntax")
        row += 2
        for line in spss_syntax.split('\n'):
            ws.cell(row=row, column=col_shift, value=line)
            row += 1

# ================================================================
# 4. Rotation Rule（旋转规则矩阵）
# ================================================================
def write_rotation_rule_to_sheet(ws, factor_target, config):
    """
    写入旋转规则矩阵，从B2开始，前三列为VariableID, Name, Lable，
    E列起为因子归属值（1显示黄色，0为空）
    """
    var_labels = config.get('factor_var_labels', {})
    var_ids = config.get('factor_var_ids', [])
    factor_cols = factor_target.columns.tolist()

    # 表头（第2行）
    header_row = 2
    ws.cell(row=header_row, column=2, value="VariableID")
    ws.cell(row=header_row, column=3, value="Name")
    ws.cell(row=header_row, column=4, value="Lable")
    for j, col_name in enumerate(factor_cols):
        ws.cell(row=header_row, column=5 + j, value=col_name)

    # 数据（从第3行开始）
    start_row = 3
    for i, var_name in enumerate(factor_target.index):
        row = start_row + i
        var_id = var_ids[i] if i < len(var_ids) else i+1
        ws.cell(row=row, column=2, value=var_id)
        ws.cell(row=row, column=3, value=var_name)
        label = var_labels.get(var_name, var_name)
        ws.cell(row=row, column=4, value=label)
        for j, col_name in enumerate(factor_cols):
            val = factor_target.loc[var_name, col_name]
            cell = ws.cell(row=row, column=5 + j)
            if pd.isna(val) or val == 0:
                cell.value = ""
                cell.fill = PatternFill(fill_type='none')
            else:
                cell.value = 1
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ================================================================
# 5. Regression Results（回归结果）
# ================================================================
def write_reg_results_to_sheet(ws, reg_results):
    """写入回归结果（模型摘要和系数表）"""
    current_row = 1
    all_models = []
    for model_name, res_dict in reg_results.items():
        for dep, res in res_dict.items():
            all_models.append((dep, res))
    if not all_models:
        return
    model_counter = 1
    for dep, res in all_models:
        r2 = res.get('R_squared', np.nan)
        adj_r2 = res.get('adj_R_squared', np.nan)
        r = res.get('R', np.nan)
        std_err = res.get('std_error', np.nan)
        coeff = res.get('coeff', pd.Series())
        bse = res.get('bse', pd.Series())
        beta = res.get('beta', pd.Series())
        tvals = res.get('tvalues', pd.Series())
        pvals = res.get('pvalues', pd.Series())
        ci_lower = res.get('ci_lower', pd.Series())
        ci_upper = res.get('ci_upper', pd.Series())
        var_names = res.get('var_names', list(coeff.index))
        if not var_names:
            var_names = list(coeff.index)

        # 因变量标题
        ws.cell(row=current_row, column=1, value=f'Dependent Variable: {dep}')
        current_row += 1
        # 模型摘要表头
        ws.cell(row=current_row, column=1, value='Model Summary')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='R')
        ws.cell(row=current_row, column=3, value='R Square')
        ws.cell(row=current_row, column=4, value='Adjusted R Square')
        ws.cell(row=current_row, column=5, value='Std. Error of the Estimate')
        current_row += 1
        ws.cell(row=current_row, column=1, value=model_counter)
        _write_num(ws.cell(row=current_row, column=2), r)
        _write_num(ws.cell(row=current_row, column=3), r2)
        _write_num(ws.cell(row=current_row, column=4), adj_r2)
        _write_num(ws.cell(row=current_row, column=5), std_err)
        current_row += 1

        # 预测变量说明
        predictors = [v for v in var_names if v != 'const']
        pred_str = '(Constant), ' + ', '.join(predictors)
        ws.cell(row=current_row, column=1, value=f'a. Predictors: {pred_str}')
        current_row += 2

        # 系数表头
        ws.cell(row=current_row, column=1, value='Coefficientsa')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='Unstandardized Coefficients')
        ws.cell(row=current_row, column=4, value='')
        ws.cell(row=current_row, column=5, value='Standardized Coefficients')
        ws.cell(row=current_row, column=6, value='t')
        ws.cell(row=current_row, column=7, value='Sig.')
        ws.cell(row=current_row, column=8, value='95.0% Confidence Interval for B')
        ws.cell(row=current_row, column=9, value='')
        current_row += 1
        ws.cell(row=current_row, column=1, value='')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='B')
        ws.cell(row=current_row, column=4, value='Std. Error')
        ws.cell(row=current_row, column=5, value='Beta')
        ws.cell(row=current_row, column=6, value='')
        ws.cell(row=current_row, column=7, value='')
        ws.cell(row=current_row, column=8, value='Lower Bound')
        ws.cell(row=current_row, column=9, value='Upper Bound')
        current_row += 1

        # 系数行
        for i, var in enumerate(var_names):
            if i == 0:
                ws.cell(row=current_row, column=1, value=model_counter)
            else:
                ws.cell(row=current_row, column=1, value='')
            display_name = 'constant' if var == 'const' else var
            ws.cell(row=current_row, column=2, value=display_name)
            _write_num(ws.cell(row=current_row, column=3), coeff.get(var, np.nan))
            _write_num(ws.cell(row=current_row, column=4), bse.get(var, np.nan))
            beta_val = beta.get(var, np.nan)
            ws.cell(row=current_row, column=5, value=beta_val if not pd.isna(beta_val) else '')
            _write_num(ws.cell(row=current_row, column=6), tvals.get(var, np.nan))
            _write_num(ws.cell(row=current_row, column=7), pvals.get(var, np.nan))
            _write_num(ws.cell(row=current_row, column=8), ci_lower.get(var, np.nan))
            _write_num(ws.cell(row=current_row, column=9), ci_upper.get(var, np.nan))
            current_row += 1

        ws.cell(row=current_row, column=2, value=f'a. Dependent Variable: {dep}')
        current_row += 2
        model_counter += 1

    # 调整列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)


def write_reg_descriptives_to_sheet(ws, reg_descriptives_info):
    """写入每个回归的描述统计（reg后Descriptives），每个回归生成三张表：原始Y、重编码Y、X"""
    from .descriptives import get_descriptives

    current_row = 1
    for info in reg_descriptives_info:
        reg_num = info['reg_num']
        filter_desc = info['filter_desc']
        df_used = info['df_used']
        y_orig_vars = info['y_orig_vars']
        y_recoded_vars = info['y_recoded_vars']
        x_vars = info['x_vars']

        # 回归标题
        title = f"Regression {reg_num}"
        if filter_desc:
            title += f" (filter: {filter_desc})"
        ws.cell(row=current_row, column=1, value=title)
        current_row += 1

        # 辅助函数：输出一个描述统计表
        def write_table(title_text, var_list):
            nonlocal current_row
            if not var_list:
                return
            # 表标题
            ws.cell(row=current_row, column=1, value=title_text)
            current_row += 1
            # 获取描述统计
            desc_df = get_descriptives(df_used, var_list, weight_series=None)
            if desc_df.empty:
                current_row += 1
                return
            # 表头
            headers = ['Variable', 'N', 'Min', 'Max', 'Mean', 'Std']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=h)
            current_row += 1
            # 数据行
            for var in var_list:
                if var not in desc_df.index:
                    continue
                row_data = desc_df.loc[var]
                ws.cell(row=current_row, column=1, value=var)
                for col_idx, field in enumerate(['N', 'Min', 'Max', 'Mean', 'Std'], start=2):
                    _write_num(ws.cell(row=current_row, column=col_idx), row_data[field])
                current_row += 1
            current_row += 1  # 表后空一行

        # 输出三张表：原始Y、重编码Y、X
        write_table("Original Y Variables", y_orig_vars)
        write_table("Recoded Y Variables", y_recoded_vars)
        write_table("X Variables (Recoded/ Merged)", x_vars)
        write_table("Recoded X Variables", info.get('recoded_x_vars', []))


        # 每个回归后额外空一行（用于分隔不同回归）
        current_row += 1
# ================================================================
# 6. Seen vs No Seen（见过 vs 没见过）
# ================================================================
def write_seen_vs_noseen_to_sheet(ws, qd_stats, qd_desc, qd_means, channel_means_dict, diff_table, listwise_n=None):
    """写入 Seen vs No Seen 分析结果，包含描述统计、均值对比和差值表及条件格式"""
    current_row = 1

    # ---- 描述统计 ----
    ws.cell(row=current_row, column=1, value='Descriptive Statistics')
    current_row += 1
    headers = ['', 'N', 'Minimum', 'Maximum', 'Mean', 'Std. Deviation']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col, value=h)
    current_row += 1

    if not qd_stats.empty:
        row_data = qd_stats.iloc[0]
        ws.cell(row=current_row, column=1, value='qd')
        _write_num(ws.cell(row=current_row, column=2), row_data['count'])
        _write_num(ws.cell(row=current_row, column=3), row_data['min'])
        _write_num(ws.cell(row=current_row, column=4), row_data['max'])
        _write_num(ws.cell(row=current_row, column=5), row_data['mean'])
        _write_num(ws.cell(row=current_row, column=6), row_data['std'])
    current_row += 1

    if listwise_n is not None and not pd.isna(listwise_n):
        ws.cell(row=current_row, column=1, value='Valid N (listwise)')
        _write_num(ws.cell(row=current_row, column=2), listwise_n)
        current_row += 1
    current_row += 1

    # ---- 均值对比表 ----
    channel_names = list(channel_means_dict.keys()) if channel_means_dict else []
    mean_headers = ['']
    if not qd_means.empty and 0 in qd_means.columns and 1 in qd_means.columns:
        mean_headers.append('qd=0')
        mean_headers.append('qd=1')
    else:
        mean_headers.append('qd=0')
        mean_headers.append('qd=1')
    for ch in channel_names:
        mean_headers.append(f'{ch} (No)')
        mean_headers.append(f'{ch} (Yes)')
    for col, h in enumerate(mean_headers, start=1):
        ws.cell(row=current_row, column=col, value=h)
    current_row += 1

    y_vars = qd_means.index.tolist() if not qd_means.empty else diff_table.index.tolist()
    for y_var in y_vars:
        ws.cell(row=current_row, column=1, value=y_var)
        col = 2
        if not qd_means.empty and y_var in qd_means.index:
            _write_num(ws.cell(row=current_row, column=col), qd_means.loc[y_var, 0])
            _write_num(ws.cell(row=current_row, column=col+1), qd_means.loc[y_var, 1])
        else:
            ws.cell(row=current_row, column=col, value='')
            ws.cell(row=current_row, column=col+1, value='')
        col += 2
        for ch in channel_names:
            gm = channel_means_dict.get(ch)
            if gm is not None and y_var in gm.columns:
                _write_num(ws.cell(row=current_row, column=col), gm.loc[0, y_var])
                _write_num(ws.cell(row=current_row, column=col+1), gm.loc[1, y_var])
            else:
                ws.cell(row=current_row, column=col, value='')
                ws.cell(row=current_row, column=col+1, value='')
            col += 2
        current_row += 1
    current_row += 1

    # ---- 差值表（带条件格式） ----
    diff_start_row = current_row
    ws.cell(row=current_row, column=1, value='')
    # qd差值
    if not qd_means.empty and 0 in qd_means.columns and 1 in qd_means.columns:
        diff_series = qd_means[1] - qd_means[0]
        ws.cell(row=current_row, column=3, value='qd (Diff)')
        for i, y_var in enumerate(y_vars):
            if y_var in diff_series.index:
                val = diff_series[y_var]
                _write_percent(ws.cell(row=current_row + 1 + i, column=3), val)
            ws.cell(row=current_row + 1 + i, column=1, value=y_var)
    # 各渠道差值
    col_offset = 5
    for ch in channel_names:
        gm = channel_means_dict.get(ch)
        if gm is not None and 0 in gm.index and 1 in gm.index:
            diff_series = gm.loc[1] - gm.loc[0]
            ws.cell(row=current_row, column=col_offset, value=f'{ch} (Diff)')
            for i, y_var in enumerate(y_vars):
                if y_var in diff_series.index:
                    val = diff_series[y_var]
                    _write_percent(ws.cell(row=current_row + 1 + i, column=col_offset), val)
        col_offset += 2

    diff_end_row = current_row + len(y_vars)

    # 条件格式
    for col_idx in range(3, col_offset, 2):
        col_letter = get_column_letter(col_idx)
        range_str = f'{col_letter}{diff_start_row + 1}:{col_letter}{diff_end_row}'
        try:
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
            ws.conditional_formatting.add(range_str, rule)
        except Exception:
            pass

    # 调整列宽
    for col in ws.columns:
        col_idx = col[0].column
        max_len = 0
        for cell in col:
            if cell.value is not None:
                if isinstance(cell.value, float) and cell.number_format == '0.00%':
                    max_len = max(max_len, 7)
                else:
                    max_len = max(max_len, len(str(cell.value)))
        if col_idx >= 2:
            width = 10
        else:
            width = min(max_len + 2, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
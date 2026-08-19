import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from .common_writers import (
    write_descriptives_to_sheet,
    write_sample_check_to_sheet,
    write_cfa_results_to_sheet,
    write_rotation_rule_to_sheet,
    write_reg_results_to_sheet,
    write_seen_vs_noseen_to_sheet,
    write_reg_descriptives_to_sheet  
)

def save_results(output_path, desc_list, loadings, score_coeff,
                 factor_corr, reg_results, qd_stats=None, qd_desc=None,
                 qd_means=None, channel_means_dict=None, diff_table=None,
                 listwise_n=None, country='', fa_extra=None, kpi_check_df=None,
                 channel_check_df=None, sample_check_success=False,
                 factor_target=None, config=None,
                 reg_descriptives_info=None, has_filtered_reg=False,sample_check_source=None):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Descriptives
        if desc_list:
            ws = writer.book.create_sheet('Descriptives')
            write_descriptives_to_sheet(ws, desc_list)

        # Sample Check
        ws = writer.book.create_sheet('Sample_Check')
        write_sample_check_to_sheet(ws, kpi_check_df, channel_check_df, sample_check_success, source=sample_check_source)

        # CFA Results
        if fa_extra is not None:
            ws = writer.book.create_sheet('CFA_Results')
            write_cfa_results_to_sheet(ws, fa_extra)

        # Factor Correlation
        if factor_corr is not None:
            write_factor_corr(writer, factor_corr, fa_extra)

        # Rotation Rule
        if factor_target is not None:
            ws = writer.book.create_sheet('RotationRule')
            write_rotation_rule_to_sheet(ws, factor_target, config)

        # Regression Results
        if reg_results:
            ws = writer.book.create_sheet('reg_Results')
            write_reg_results_to_sheet(ws, reg_results)

        # Seen vs No Seen
        if qd_stats is not None and qd_means is not None:
            ws = writer.book.create_sheet('seen vs no seen')
            write_seen_vs_noseen_to_sheet(ws, qd_stats, qd_desc, qd_means,
                                          channel_means_dict, diff_table, listwise_n)

        # ---- reg后Descriptives ----
        if reg_descriptives_info:
            ws = writer.book.create_sheet('reg后Descriptives')
            write_reg_descriptives_to_sheet(ws, reg_descriptives_info)

def write_factor_corr(writer, factor_corr, fa_extra):
    # 保持不变
    ws = writer.book.create_sheet('FactorCorr')
    n_factors = fa_extra.get('n_factors', factor_corr.shape[0])
    factor_labels = [f'Factor{i+1}' for i in range(n_factors)]
    df_corr = pd.DataFrame(factor_corr, index=factor_labels, columns=factor_labels)
    df_corr.to_excel(writer, sheet_name='FactorCorr', startrow=0, startcol=0)
    ws = writer.sheets['FactorCorr']
    for r in range(1, n_factors + 1):
        for c in range(1, n_factors + 1):
            cell = ws.cell(row=r + 1, column=c + 1)
            val = cell.value
            if val is not None:
                if abs(val) < 1e-9:
                    cell.number_format = '0.000'
                else:
                    if abs(val - round(val)) < 1e-12:
                        cell.number_format = '0'
                    else:
                        cell.number_format = '0.000'
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 12)
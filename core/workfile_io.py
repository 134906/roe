# workfile_io.py
import openpyxl
import tempfile
import shutil
import os
from .common_writers import (
    write_descriptives_to_sheet,
    write_sample_check_to_sheet,
    write_cfa_results_to_sheet,
    write_rotation_rule_to_sheet,
    write_reg_results_to_sheet,
    write_seen_vs_noseen_to_sheet,
    write_reg_descriptives_to_sheet
)

def fill_workfile(workfile_path, output_workfile_path, config, desc_list, fa_extra,
                  reg_results, qd_stats, qd_desc, qd_means, channel_means_dict,
                  diff_table, listwise_n, kpi_check_df, channel_check_df,
                  sample_check_success, factor_corr, name_label_map=None, log_callback=print,
                  reg_descriptives_info=None, has_filtered_reg=False,sample_check_source=None):
    log_callback("开始填充 workfile 模板...")
    
    # ---- 处理文件被占用问题 ----
    tmp_path = None
    try:
        wb = openpyxl.load_workbook(workfile_path)
    except PermissionError:
        log_callback(f"模板文件 {workfile_path} 被占用，将复制到临时文件后操作...")
        # 创建临时文件（保留 .xlsx 扩展名）
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        shutil.copy2(workfile_path, tmp_path)
        wb = openpyxl.load_workbook(tmp_path)
    
    # ---- 后续填充操作 ----
    def get_or_create_sheet(wb, name):
        return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

    # 1. 2A 回归syntax
    reg_syntax_sheet = get_or_create_sheet(wb, '2A 回归syntax')
    row = 2
    if reg_descriptives_info:
        for info in reg_descriptives_info:
            dep_vars = info.get('dep_vars', [])
            ind_vars = info.get('ind_vars', [])
            max_len = max(len(dep_vars), len(ind_vars))
            for i in range(max_len):
                if i < len(ind_vars):
                    reg_syntax_sheet.cell(row=row, column=1, value=ind_vars[i])
                if i < len(dep_vars):
                    reg_syntax_sheet.cell(row=row, column=2, value=dep_vars[i])
                row += 1
            row += 3
    else:
        for idx, reg_cfg in enumerate(config.get('reg_configs', [])):
            dep_orig = reg_cfg.get('dep', [])
            ind_orig = reg_cfg.get('ind', [])
            max_len = max(len(dep_orig), len(ind_orig))
            for i in range(max_len):
                if i < len(ind_orig):
                    reg_syntax_sheet.cell(row=row, column=1, value=ind_orig[i])
                if i < len(dep_orig):
                    reg_syntax_sheet.cell(row=row, column=2, value=dep_orig[i])
                row += 1
            row += 3
    log_callback("  2A 回归syntax 填充完成")

    # 2. 2A Regression
    ws = get_or_create_sheet(wb, '2A Regression')
    write_reg_results_to_sheet(ws, reg_results)
    log_callback("  2A Regression 填充完成")

    # 3. 2B RotationRule
    if config.get('factor_target') is not None:
        ws = get_or_create_sheet(wb, '2B RotationRule')
        write_rotation_rule_to_sheet(ws, config['factor_target'], config)
        log_callback("  2B RotationRule 填充完成")

    # 4. 2B FL
    if fa_extra is not None:
        ws = get_or_create_sheet(wb, '2B FL')
        write_cfa_results_to_sheet(ws, fa_extra)
        log_callback("  2B FL 填充完成")

    # 5. 2C 均值
    ws = get_or_create_sheet(wb, '2C 均值')
    if has_filtered_reg and reg_descriptives_info:
        filtered_infos = [info for info in reg_descriptives_info if info['filter_desc'] is not None]
        write_reg_descriptives_to_sheet(ws, filtered_infos)
    else:
        if desc_list:
            write_descriptives_to_sheet(ws, desc_list)
    log_callback("  2C 均值 填充完成")

    # 6. seenvsnoseen
    if qd_stats is not None and qd_means is not None:
        ws = get_or_create_sheet(wb, 'seenvsnoseen')
        write_seen_vs_noseen_to_sheet(ws, qd_stats, qd_desc, qd_means,
                                      channel_means_dict, diff_table, listwise_n)
        log_callback("  seenvsnoseen 填充完成")

    # 7. Sample_Check
    ws = get_or_create_sheet(wb, 'Sample_Check')
    write_sample_check_to_sheet(ws, kpi_check_df, channel_check_df, sample_check_success, source=sample_check_source)
    log_callback("  Sample_Check 填充完成")

    # 保存到输出路径
    wb.save(output_workfile_path)
    wb.close()
    
    # 清理临时文件
    if tmp_path and os.path.exists(tmp_path):
        try:
            os.unlink(tmp_path)
        except OSError:
            pass  # 忽略删除失败

    log_callback(f"Workfile 填充完成，保存至: {output_workfile_path}")